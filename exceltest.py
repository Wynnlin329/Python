from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment

wb = Workbook()

ws = wb.active
ws.title ='Sheet1'

wb.create_sheet('Sheet2',index=1)

ws['A1'] = '哈囉'
ws.append(['A','B','C'])
ws.append([200,300,500,'=AVERAGE(A3:C3)'])
ws.cell(row=4,column=1).value = 600
ws.cell(4,2).value = 700
ws= wb['Sheet2']
ws['A1'] ='測試'
myfont = Font(name='微軟正黑體', size=18, italic=True, color=colors.BLUE, bold=True)
ws['A1'].font = myfont

#存檔
wb.save('out.xlsx')
print('ok')


