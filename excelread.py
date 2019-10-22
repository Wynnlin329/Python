from openpyxl import load_workbook

wb = load_workbook('out.xlsx')

print(wb.sheetnames)
print(wb.active)

ws = wb['Sheet1']

print(ws['A3'].value)
print(ws.cell(row=3,column=1).value)
print(ws.cell(3,1).value)

for row in ws.rows:
    for cell in row:
        print(cell.value)

ws = wb['Sheet2']
print(ws['A1'].value)
